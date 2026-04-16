import os
import json
import random
import tempfile
import subprocess
import datetime
from flask import Flask, jsonify, request
from flask_cors import CORS
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

# Configure CORS for production
CORS(app, origins=[
    "https://mohsinfurkh.github.io",  # Your GitHub Pages URL
    "http://localhost:3000",           # Local development
    "https://127.0.0.1:3000",
    "*"  # Allow all origins for debugging (remove in production)
])

# Production configuration
app.config['DEBUG'] = False
app.config['TESTING'] = False
app.config['ENV'] = 'production'

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
QUESTIONS_FILE = os.path.join(BASE_DIR, 'Exercises', 'questions.json')
EXCEL_FILE = os.path.join(BASE_DIR, 'Student_Marks.xlsx')

def get_questions():
    if not os.path.exists(QUESTIONS_FILE):
        return []
    with open(QUESTIONS_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)

def run_code_with_input(code, standard_input):
    """Saves code to a temp file, runs it with given input, and returns (stdout, stderr)."""
    with tempfile.NamedTemporaryFile(mode='w', suffix='.py', delete=False, encoding='utf-8') as temp_script:
        temp_script.write(code)
        temp_script_path = temp_script.name

    try:
        # Try different Python executables
        python_commands = ['python3', 'python', 'python.exe']
        process = None
        last_error = None
        
        for python_cmd in python_commands:
            try:
                process = subprocess.run(
                    [python_cmd, temp_script_path],
                    input=standard_input,
                    capture_output=True,
                    text=True,
                    timeout=5
                )
                break  # Success, exit the loop
            except FileNotFoundError:
                last_error = f"{python_cmd} not found"
                continue
            except subprocess.TimeoutExpired:
                return "", "Execution Timed Out"
        
        if process is None:
            return "", f"Python not found. Tried: {', '.join(python_commands)}. Error: {last_error}"
        
        return process.stdout, process.stderr
    except Exception as e:
        return "", str(e)
    finally:
        # Clean up temp file
        if os.path.exists(temp_script_path):
            os.remove(temp_script_path)

def generate_alternative_inputs(standard_input):
    """Generate alternative input formats for flexible testing."""
    alternatives = [standard_input]  # Original format
    
    # If input contains space-separated numbers, create line-separated version
    stripped_input = standard_input.strip()
    if stripped_input and ' ' in stripped_input:
        # Split by whitespace and rejoin with newlines
        parts = stripped_input.split()
        line_separated = '\n'.join(parts) + '\n'
        alternatives.append(line_separated)
    
    return alternatives

def normalize_output(text):
    """Normalize output by stripping and handling different line endings."""
    return text.strip().replace('\r\n', '\n')

def generate_alternative_outputs(expected_output):
    """Generate alternative output formats for flexible comparison."""
    alternatives = [expected_output]  # Original format
    
    stripped_output = expected_output.strip()
    if stripped_output:
        # If output contains space-separated values, create line-separated version
        if ' ' in stripped_output and not any(c in stripped_output for c in '.!?,;:'):
            parts = stripped_output.split()
            line_separated = '\n'.join(parts) + '\n'
            alternatives.append(line_separated)
        
        # If output contains line-separated values, create space-separated version
        elif '\n' in stripped_output:
            lines = stripped_output.split('\n')
            # Only create space-separated version if all lines contain single values/words
            if all(len(line.strip().split()) == 1 for line in lines if line.strip()):
                space_separated = ' '.join(line.strip() for line in lines if line.strip()) + '\n'
                alternatives.append(space_separated)
    
    return alternatives

@app.route('/api/get_question', methods=['GET', 'POST'])
def api_get_question():
    # Get excluded question IDs from request
    excluded_ids = []
    if request.method == 'POST':
        data = request.get_json() or {}
        excluded_ids = data.get('excluded_ids', [])
    
    questions = get_questions()
    if not questions:
        return jsonify({"error": "No questions available."}), 404
    
    # Filter out excluded questions
    available_questions = [q for q in questions if q['id'] not in excluded_ids]
    
    if not available_questions:
        return jsonify({"error": "All questions have been shown. Please refresh to start a new session."}), 404
    
    selected = random.choice(available_questions)
    if selected.get("test_cases"):
        sample_tc = selected["test_cases"][0]
        sample_input = sample_tc.get("input", "")
        sample_output = sample_tc.get("expected_output", "")
    else:
        sample_input = ""
        sample_output = ""
    
    return jsonify({
        "id": selected["id"], 
        "text": selected["text"],
        "sample_input": sample_input,
        "sample_output": sample_output,
        "remaining_questions": len(available_questions) - 1
    })

@app.route('/api/reset_session', methods=['POST'])
def api_reset_session():
    """Reset the session to allow all questions to be available again."""
    return jsonify({"message": "Session reset successfully"})

@app.route('/api/evaluate', methods=['POST'])
def api_evaluate():
    data = request.json
    name = data.get('name', '').strip()
    sap_id = data.get('sap_id', '').strip()
    question_id = data.get('question_id')
    code = data.get('code', '')
    security_data = data.get('security_data', {})

    if not name or not sap_id or question_id is None or not code:
        return jsonify({"error": "Missing required fields"}), 400

    questions = get_questions()
    question = next((q for q in questions if q['id'] == question_id), None)
    
    if not question:
        return jsonify({"error": "Question not found"}), 404

    test_cases = question.get('test_cases', [])
    test_results = []  # Store detailed results for each test case
    
    if not test_cases:
        # If there are no test cases, we can't reliably grade it, but let's give 7 to be nice, or 0.
        marks = 0
    else:
        passed = 0
        total = len(test_cases)
        
        for i, tc in enumerate(test_cases, 1):
            tc_input = tc.get('input', '')
            expected_output = tc.get('expected_output', '')
            
            # Try multiple input formats
            input_formats = generate_alternative_inputs(tc_input)
            output_formats = generate_alternative_outputs(expected_output)
            test_passed = False
            actual_output = ""
            error_msg = ""
            
            for input_format in input_formats:
                stdout, stderr = run_code_with_input(code, input_format)
                
                if stderr:
                    error_msg = stderr
                    continue  # Try next input format if there's an error
                
                # Check against multiple output formats
                for output_format in output_formats:
                    if normalize_output(stdout) == normalize_output(output_format):
                        test_passed = True
                        actual_output = stdout
                        break  # Found matching output format
                
                if test_passed:
                    break  # Test passed, no need to try other input formats
            
            # Store test case result
            test_result = {
                "test_case": i,
                "input": tc_input.strip(),
                "expected": expected_output.strip(),
                "passed": test_passed
            }
            
            if test_passed:
                test_result["actual"] = actual_output.strip()
                passed += 1
            else:
                if error_msg:
                    test_result["error"] = error_msg
                else:
                    # Get the last attempted output for feedback
                    last_stdout, _ = run_code_with_input(code, tc_input)
                    test_result["actual"] = last_stdout.strip()
                    test_result["feedback"] = get_failure_feedback(expected_output.strip(), last_stdout.strip())
            
            test_results.append(test_result)

        marks = round((passed / total) * 7.0, 2)

    # Log to Excel with security data
    log_to_excel(name, sap_id, question_id, marks, security_data)

    return jsonify({
        "marks": marks, 
        "total_marks": 7, 
        "message": "Evaluation successful",
        "test_results": test_results,
        "passed_tests": passed,
        "total_tests": len(test_cases) if test_cases else 0
    })

def get_failure_feedback(expected, actual):
    """Generate helpful feedback based on the type of mismatch."""
    if not actual:
        return "No output produced. Make sure your program prints the result."
    
    # Check for common issues
    if actual.strip() == expected.strip():
        return "Output matches but may have extra whitespace or formatting issues."
    
    # Check if it's a formatting issue
    expected_parts = expected.replace('\n', ' ').split()
    actual_parts = actual.replace('\n', ' ').split()
    
    if expected_parts == actual_parts:
        return "Correct values but different formatting (space vs newline separation)."
    
    # Check if numbers are close (for floating point)
    try:
        expected_nums = [float(x) for x in expected.replace('\n', ' ').split()]
        actual_nums = [float(x) for x in actual.replace('\n', ' ').split()]
        if len(expected_nums) == len(actual_nums):
            if all(abs(e - a) < 0.001 for e, a in zip(expected_nums, actual_nums)):
                return "Correct numerical values but precision/formatting difference."
    except ValueError:
        pass
    
    return f"Expected: '{expected}', Got: '{actual}'"

def log_to_excel(name, sap_id, question_id, marks, security_data=None):
    try:
        if not os.path.exists(EXCEL_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = "Marks"
            headers = [
                "Timestamp", "Name", "SAP ID", "Question ID", "Marks Out of 7",
                "Test Duration (seconds)", "Tab Switches", "Warnings", 
                "Start Time", "End Time", "Security Violations"
            ]
            ws.append(headers)
        else:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active

        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Extract security data
        test_duration = security_data.get('test_duration', 0) if security_data else 0
        tab_switches = security_data.get('tab_switches', 0) if security_data else 0
        warnings = security_data.get('warnings', 0) if security_data else 0
        start_time = security_data.get('start_time', '') if security_data else ''
        end_time = security_data.get('end_time', '') if security_data else ''
        
        # Count security violations
        security_violations = 0
        if security_data and 'security_alerts' in security_data:
            security_violations = len([alert for alert in security_data['security_alerts'] 
                                     if alert.get('type') == 'tab_switch'])
        
        row = [
            timestamp, name, sap_id, question_id, marks,
            test_duration, tab_switches, warnings,
            start_time, end_time, security_violations
        ]
        ws.append(row)
        wb.save(EXCEL_FILE)
    except Exception as e:
        print(f"Error saving to Excel: {e}")

@app.route('/health')
def health_check():
    """Health check endpoint for Render"""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.datetime.now().isoformat(),
        'version': '1.0.0'
    })

if __name__ == '__main__':
    # Get port from environment variable for Render
    port = int(os.environ.get('PORT', 5000))
    # Run on all interfaces for Render with production settings
    app.run(host='0.0.0.0', port=port, debug=False, threaded=True)
